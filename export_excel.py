import os, sys
try:
    import openpyxl
except ImportError:
    os.system(f'"{sys.executable}" -m pip install openpyxl')
    import openpyxl

from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime

def export_logbook_to_excel(config, ledger, all_trips, output_file_path):
    wb = openpyxl.Workbook()

    # Styles
    title_font = Font(name='Calibri', size=16, bold=True, color='1F4E79')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    bold_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=11)
    small_font = Font(name='Calibri', size=10)

    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    light_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    summary_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    bottom_border = Border(bottom=Side(style='medium'))

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')

    # Count visits per client
    visit_counts = {}
    for _, legs in all_trips:
        clients_visited = set()
        for frm, to, _ in legs:
            if to not in ("Work", "Home"): clients_visited.add(to)
            if frm not in ("Work", "Home"): clients_visited.add(frm)
        for c in clients_visited:
            visit_counts[c] = visit_counts.get(c, 0) + 1
    total_visits = sum(visit_counts.values())

    # ============ SHEET 1: COVER SHEET ============
    ws1 = wb.active
    ws1.title = "Cover Sheet"
    ws1.sheet_properties.tabColor = "1F4E79"

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 20

    # Title
    ws1.merge_cells('B2:C2')
    ws1['B2'].value = "BUSINESS TRAVEL LOGBOOK"
    ws1['B2'].font = Font(name='Calibri', size=20, bold=True, color='1F4E79')
    ws1['B2'].alignment = center_align

    # Period string
    try:
        s_dt = datetime.strptime(config["start_date"], "%Y-%m-%d")
        e_dt = datetime.strptime(config["end_date"], "%Y-%m-%d")
        period_str = f"{s_dt.strftime('%d %B %Y')} – {e_dt.strftime('%d %B %Y')}"
    except Exception:
        period_str = f"{config['start_date']} – {config['end_date']}"

    ws1.merge_cells('B3:C3')
    ws1['B3'].value = period_str
    ws1['B3'].font = Font(name='Calibri', size=14, color='4472C4')
    ws1['B3'].alignment = center_align

    open_odo = config["opening_odo"]
    close_odo = config["closing_odo"]
    biz_km = config["target_business_km"]
    total_km = close_odo - open_odo
    private_km = total_km - biz_km

    summary_data = [
        ("", ""),
        ("Vehicle Summary", ""),
        ("Opening Odometer:", f"{open_odo:,} km"),
        ("Closing Odometer:", f"{close_odo:,} km"),
        ("", ""),
        ("Kilometre Summary", ""),
        ("Total Vehicle km:", f"{total_km:,} km"),
        ("Business km:", f"{biz_km:,} km"),
        ("Private km:", f"{private_km:,} km"),
    ]

    row = 5
    for label, value in summary_data:
        ws1.cell(row=row, column=2, value=label)
        ws1.cell(row=row, column=3, value=value)
        if "Summary" in label:
            ws1.cell(row=row, column=2).font = bold_font
            ws1.cell(row=row, column=2).border = bottom_border
            ws1.cell(row=row, column=3).border = bottom_border
        elif label:
            ws1.cell(row=row, column=2).font = normal_font
            ws1.cell(row=row, column=3).font = normal_font
            ws1.cell(row=row, column=3).alignment = right_align
        row += 1

    # Client visit summary
    row += 1
    ws1.cell(row=row, column=2, value="Client Visit Summary").font = bold_font
    ws1.cell(row=row, column=2).border = bottom_border
    ws1.cell(row=row, column=3).border = bottom_border
    row += 1

    for col, header in enumerate(["Client", "Visits"], 2):
        cell = ws1.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    for i, c in enumerate(config["clients"]):
        fill = light_fill if i % 2 == 0 else white_fill
        c_name = c["name"]
        cell = ws1.cell(row=row, column=2, value=c_name)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border

        cell = ws1.cell(row=row, column=3, value=visit_counts.get(c_name, 0))
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center_align
        row += 1

    # Total row
    cell = ws1.cell(row=row, column=2, value="TOTAL")
    cell.font = bold_font
    cell.fill = summary_fill
    cell.border = thin_border
    cell = ws1.cell(row=row, column=3, value=total_visits)
    cell.font = bold_font
    cell.fill = summary_fill
    cell.border = thin_border
    cell.alignment = center_align
    row += 1

    row += 1

    # ============ SHEET 2: BUSINESS LOGBOOK ============
    ws2 = wb.create_sheet("Business Logbook")
    ws2.sheet_properties.tabColor = "4472C4"

    col_widths = {'A': 14, 'B': 28, 'C': 28, 'D': 16, 'E': 16, 'F': 14}
    for col, width in col_widths.items():
        ws2.column_dimensions[col].width = width

    ws2.merge_cells('A1:F1')
    ws2['A1'].value = f"BUSINESS TRAVEL LOGBOOK — {period_str}"
    ws2['A1'].font = title_font
    ws2['A1'].alignment = center_align

    headers = ["Date", "From", "To", "Opening Odo", "Closing Odo", "Business km"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    num_fmt = '#,##0' if config.get("round_numbers_only", True) else '#,##0.#'

    row = 4
    for entry_date, frm, to, o_odo, c_odo, b_km in ledger:
        fill = light_fill if (row - 4) % 4 < 2 else white_fill

        cell = ws2.cell(row=row, column=1, value=entry_date)
        cell.number_format = 'DD/MM/YYYY'
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = center_align

        cell = ws2.cell(row=row, column=2, value=frm)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = left_align

        cell = ws2.cell(row=row, column=3, value=to)
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = left_align

        cell = ws2.cell(row=row, column=4, value=o_odo)
        cell.number_format = num_fmt
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = right_align

        cell = ws2.cell(row=row, column=5, value=c_odo)
        cell.number_format = num_fmt
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = right_align

        cell = ws2.cell(row=row, column=6, value=b_km)
        cell.number_format = num_fmt
        cell.font = normal_font
        cell.fill = fill
        cell.border = thin_border
        cell.alignment = right_align

        row += 1

    row += 1
    cell = ws2.cell(row=row, column=5, value="TOTAL BUSINESS KM:")
    cell.font = bold_font
    cell.alignment = right_align

    cell = ws2.cell(row=row, column=6, value=f"=SUM(F4:F{row-2})")
    cell.font = bold_font
    cell.fill = summary_fill
    cell.border = thin_border
    cell.alignment = right_align
    cell.number_format = num_fmt

    ws2.freeze_panes = 'A4'

    os.makedirs(os.path.dirname(os.path.abspath(output_file_path)), exist_ok=True)
    wb.save(output_file_path)
    return output_file_path
